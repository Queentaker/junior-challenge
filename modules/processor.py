import shutil
import pandas as pd
import openpyxl as xl
import os.path
from datetime import datetime

import modules.constants as const
from modules.identifier import generate_identifier


def run(file_name: str):

    def create_excels(csv: str, target_folder: str):
        assert os.path.exists(target_folder)
        df=pd.read_csv(csv)
        for index,row in df.iterrows():
            identifier=generate_identifier(8)
            file_name= f"{identifier}" + ".xlsx"

            file_path=os.path.join(target_folder, file_name)
            shutil.copy(const.template_path, file_path)

            wb = xl.load_workbook(file_path)
            ws =wb.worksheets[0]

            ws[const.identifier_cell].value=identifier
            ws[const.firstname_cell].value=row["firstname"]
            ws[const.lastname_cell].value=row["lastname"]
            ws[const.matriculation_cell].value=row["matriculation_number"]

            wb.save(file_path)

    def generate_folder_name():
        folder_name = "Folder_Excelfiles_" + datetime.now().now().strftime("%Y-%m-%d_%H_%M_%S")
        return folder_name

    def create_folder(folder_name: str):
        assert not os.path.exists(folder_name)
        os.makedirs(folder_name, exist_ok=True)

    if not os.path.exists(file_name):
        print("File path doesn't exist")
    else:
        print(">> Processing...")

        folder_name = generate_folder_name()
        create_folder(folder_name)
        create_excels(file_name, folder_name)

        # ---- YOUR CODE ENDS HERE ----
        print(">> Finished processing.")

